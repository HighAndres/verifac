"""Pruebas del export mensual a Excel."""
import io
from datetime import datetime, timezone

import openpyxl

from app.models.factura import Factura
from app.services.export_excel import generar_excel_mes
from tests.factories import add_montos, add_profesor, build_cfdi


def _factura(db, cfdi, estado, motivo_rechazo=None):
    f = Factura(
        uuid_cfdi=cfdi.uuid, rfc_emisor=cfdi.rfc_emisor, nombre_emisor=cfdi.nombre_emisor,
        regimen_emisor=cfdi.regimen_fiscal, nombre_receptor=cfdi.nombre_receptor,
        moneda=cfdi.moneda, fecha_emision=cfdi.fecha, subtotal=cfdi.subtotal,
        iva_trasladado=cfdi.iva_trasladado, iva_retenido=cfdi.iva_retenido,
        isr_retenido=cfdi.isr_retenido, total=cfdi.total,
        clave_servicio="90141702", clave_unidad="E48", uso_cfdi=cfdi.uso_cfdi,
        forma_pago=cfdi.forma_pago, metodo_pago=cfdi.metodo_pago, estado=estado,
        motivo_rechazo=motivo_rechazo,
        origen="xml", pdf_cotejo="sin_pdf", fecha_validacion=datetime.now(timezone.utc),
    )
    db.add(f)
    db.flush()
    return f


def _abrir(contenido: bytes):
    return openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)


def test_export_tiene_tres_hojas_y_estatus(db):
    p1 = add_profesor(db, rfc="EXP010101AA1", nombre="EXPORT UNO", regimen="612")
    p2 = add_profesor(db, rfc="EXP010101BB2", nombre="EXPORT DOS", regimen="612")
    p3 = add_profesor(db, rfc="EXP010101CC3", nombre="EXPORT TRES", regimen="612")
    c1 = build_cfdi(rfc_emisor=p1.rfc, nombre_emisor=p1.nombre,
                    fecha=datetime(2099, 3, 10), uuid="EXP-1")
    c2 = build_cfdi(rfc_emisor=p2.rfc, nombre_emisor=p2.nombre,
                    fecha=datetime(2099, 3, 12), uuid="EXP-2")
    c3 = build_cfdi(rfc_emisor=p3.rfc, nombre_emisor=p3.nombre,
                    fecha=datetime(2099, 3, 15), uuid="EXP-3")
    add_montos(db, p1, c1)
    add_montos(db, p2, c2)
    _factura(db, c1, "aprobada")          # p1 aprobada; p2 sin factura
    _factura(db, c3, "rechazada", motivo_rechazo="Campos con error: Uso CFDI")

    wb = _abrir(generar_excel_mes(db, 3, 2099))
    assert wb.sheetnames == ["Resumen conciliación", "Base BBVA", "Aprobadas y rechazadas"]

    ws = wb["Resumen conciliación"]
    filas = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
    assert filas["EXPORT UNO"][9] == "Aprobada"
    assert filas["EXPORT UNO"][10] == "EXP-1"          # UUID
    assert filas["EXPORT UNO"][13] == 0                # diferencia total
    assert filas["EXPORT DOS"][9] == "Sin factura"
    assert filas["EXPORT DOS"][12] is None             # sin total facturado

    ws2 = wb["Base BBVA"]
    aprobadas = list(ws2.iter_rows(min_row=2, values_only=True))
    assert len(aprobadas) == 1                          # solo la aprobada
    fila = aprobadas[0]
    assert fila[7] == "EXPORT UNO"                      # nombre emisor
    assert fila[18] == float(c1.total)                  # total

    ws3 = wb["Aprobadas y rechazadas"]
    todas = {r[0]: r for r in ws3.iter_rows(min_row=2, values_only=True)}
    assert set(todas) == {"EXPORT UNO", "EXPORT TRES"}   # p2 nunca subió factura
    assert todas["EXPORT UNO"][2] == "Aprobada"
    assert todas["EXPORT UNO"][3] is None                # sin motivo, se aprobó
    assert todas["EXPORT TRES"][2] == "Rechazada"
    assert todas["EXPORT TRES"][3] == "Campos con error: Uso CFDI"


def test_export_mes_vacio(db):
    wb = _abrir(generar_excel_mes(db, 4, 2099))
    ws = wb["Resumen conciliación"]
    assert list(ws.iter_rows(min_row=2, values_only=True)) == []

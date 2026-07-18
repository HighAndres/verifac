"""
Exportación del mes conciliado a Excel (.xlsx).

Hoja 1 "Resumen conciliación": una fila por entrada del layout de montos, con los
montos esperados, el estatus de su factura y la diferencia contra lo facturado.
Hoja 2 "Base BBVA": las facturas APROBADAS del mes en las columnas del formato
"Ejemplo Base BBVA" (el excel que se envía para pago).
"""
import io
from datetime import datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import extract
from sqlalchemy.orm import Session

from app.models.factura import Factura
from app.models.monto_mensual import MontoMensual
from app.models.profesor import Profesor

REGIMEN_DESC = {
    "612": "Personas Físicas con Actividades Empresariales y Profesionales",
    "626": "Régimen Simplificado de Confianza",
    "603": "Personas Morales con Fines no Lucrativos",
}
UNIDAD_DESC = {"E48": "Unidad de servicio", "ACT": "Actividad"}

_MONEY = "#,##0.00"


def _hoja_encabezado(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1E3A8A")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
    ws.freeze_panes = "A2"


def _autoancho(ws, minimo: int = 10, maximo: int = 40) -> None:
    for col in ws.columns:
        ancho = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        letra = get_column_letter(col[0].column)
        ws.column_dimensions[letra].width = min(max(ancho + 2, minimo), maximo)


def generar_excel_mes(db: Session, mes: int, anio: int) -> bytes:
    del_mes = [
        extract("month", Factura.fecha_emision) == mes,
        extract("year", Factura.fecha_emision) == anio,
    ]
    facturas = db.query(Factura).filter(*del_mes).all()
    # Por RFC nos quedamos con la aprobada si existe; si no, la más reciente.
    por_rfc: dict[str, Factura] = {}
    for f in facturas:
        actual = por_rfc.get(f.rfc_emisor)
        if actual is None or (f.estado == "aprobada" and actual.estado != "aprobada"):
            por_rfc[f.rfc_emisor] = f

    montos = (
        db.query(MontoMensual)
        .filter(MontoMensual.mes == mes, MontoMensual.anio == anio)
        .order_by(MontoMensual.nombre_layout)
        .all()
    )

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen conciliación ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen conciliación"
    _hoja_encabezado(ws, [
        "Nombre", "RFC", "Régimen", "Categoría",
        "Subtotal esperado", "IVA esperado", "IVA Ret esperado",
        "ISR Ret esperado", "Total esperado",
        "Estatus", "UUID factura", "Fecha emisión", "Total facturado", "Diferencia",
    ])

    for m in montos:
        prof = db.get(Profesor, m.profesor_id) if m.profesor_id else None
        rfc = prof.rfc if prof else (m.rfc_emisor or "")
        f = por_rfc.get(rfc) if rfc else None

        if prof is None:
            estatus = "Sin match en catálogo"
        elif f is None:
            estatus = "Sin factura"
        elif f.estado == "aprobada":
            estatus = "Aprobada"
        else:
            estatus = f.estado.capitalize()

        total_fact = Decimal(str(f.total)) if f is not None and f.total is not None else None
        diferencia = (
            (total_fact - Decimal(str(m.total))) if total_fact is not None else None
        )
        ws.append([
            m.nombre_layout, rfc, m.regimen_fiscal, m.categoria,
            float(m.subtotal), float(m.iva_trasladado), float(m.iva_retenido),
            float(m.isr_retenido), float(m.total),
            estatus,
            f.uuid_cfdi if f else None,
            f.fecha_emision.strftime("%Y-%m-%d") if f and f.fecha_emision else None,
            float(total_fact) if total_fact is not None else None,
            float(diferencia) if diferencia is not None else None,
        ])

    for fila in ws.iter_rows(min_row=2):
        for c in fila[4:9] + fila[12:14]:
            c.number_format = _MONEY
    _autoancho(ws)

    # ── Hoja 2: Base BBVA (facturas aprobadas del mes) ───────────────────────
    ws2 = wb.create_sheet("Base BBVA")
    _hoja_encabezado(ws2, [
        "Categoría", "Fecha de emisión", "Uso CFDI", "Clave prod/serv",
        "Concepto de servicio", "Clave régimen emisor", "Régimen emisor",
        "Nombre emisor", "Clave régimen receptor", "Régimen receptor",
        "Nombre receptor", "Unidad", "Descripción unidad", "Cantidad",
        "Subtotal", "IVA Trasladado", "IVA Retenido", "ISR retenido", "Total",
        "Moneda", "Forma de pago", "Método de pago",
    ])

    categoria_por_rfc = {
        (db.get(Profesor, m.profesor_id).rfc if m.profesor_id else m.rfc_emisor): m.categoria
        for m in montos
        if m.profesor_id or m.rfc_emisor
    }
    aprobadas = sorted(
        (f for f in facturas if f.estado == "aprobada"),
        key=lambda f: (f.nombre_emisor or ""),
    )
    for f in aprobadas:
        ws2.append([
            categoria_por_rfc.get(f.rfc_emisor),
            f.fecha_emision.strftime("%Y-%m-%d") if f.fecha_emision else None,
            f.uso_cfdi, f.clave_servicio, f.descripcion_concepto,
            f.regimen_emisor, REGIMEN_DESC.get(f.regimen_emisor or "", ""),
            f.nombre_emisor, None, None, f.nombre_receptor,
            f.clave_unidad, UNIDAD_DESC.get(f.clave_unidad or "", ""), 1,
            float(f.subtotal or 0), float(f.iva_trasladado or 0),
            float(f.iva_retenido or 0), float(f.isr_retenido or 0), float(f.total or 0),
            f.moneda, f.forma_pago, f.metodo_pago,
        ])

    for fila in ws2.iter_rows(min_row=2):
        for c in fila[14:19]:
            c.number_format = _MONEY
    _autoancho(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def nombre_archivo(mes: int, anio: int) -> str:
    return f"Verifac_conciliacion_{anio}-{mes:02d}.xlsx"

"""
Parser para el layout de montos mensuales (Ejemplo Base BBVA Montos.xlsx).
Columnas esperadas (en orden): Categoría, Clave régimen emisor, Nombre emisor,
Subtotal, IVA Trasladado, IVA Retenido, ISR retenido, Total.

Columnas opcionales, detectadas por encabezado en cualquier posición extra:
  - RFC  → llave preferida de emparejamiento con el catálogo de profesores.
  - Mes  → número 1-12 o nombre del mes (enero…diciembre).
  - Año  → año de 4 dígitos.
Si Mes/Año están presentes, el endpoint valida que coincidan con el periodo
seleccionado para evitar cargar el layout en el mes equivocado.
"""
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl

COLUMNAS = ("categoria", "regimen_fiscal", "nombre_emisor",
            "subtotal", "iva_trasladado", "iva_retenido", "isr_retenido", "total")

_MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


@dataclass
class FilaMonto:
    fila: int
    categoria: Optional[str]
    regimen_fiscal: str
    nombre_emisor: str
    rfc_emisor: Optional[str]
    mes: Optional[int]
    anio: Optional[int]
    subtotal: Decimal
    iva_trasladado: Decimal
    iva_retenido: Decimal
    isr_retenido: Decimal
    total: Decimal


def normalizar_nombre(valor: Optional[str]) -> str:
    """Normaliza un nombre para emparejar sin depender de acentos, espacios ni mayúsculas."""
    if not valor:
        return ""
    texto = unicodedata.normalize("NFKD", str(valor))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return " ".join(texto.upper().split())


def _dec(val) -> Decimal:
    if val is None:
        return Decimal("0")
    try:
        return Decimal(str(val)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0")


def _parse_mes(val) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        n = int(float(s))
        return n if 1 <= n <= 12 else None
    except ValueError:
        return _MESES.get(normalizar_nombre(s))


def _parse_anio(val) -> Optional[int]:
    if val is None:
        return None
    try:
        n = int(float(str(val).strip()))
        return n if 2000 <= n <= 2100 else None
    except ValueError:
        return None


def _buscar_columna(ws, *claves: str) -> Optional[int]:
    """Índice (0-based) de la primera columna cuyo encabezado contenga alguna de las claves."""
    encabezado = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not encabezado:
        return None
    for idx, celda in enumerate(encabezado):
        if not celda:
            continue
        texto = normalizar_nombre(celda)
        if any(clave in texto for clave in claves):
            return idx
    return None


def _celda(row, idx: Optional[int]):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def parsear_excel_montos(contenido: bytes) -> list[FilaMonto]:
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    ws = wb.active

    col_rfc = _buscar_columna(ws, "RFC")
    col_mes = _buscar_columna(ws, "MES")
    col_anio = _buscar_columna(ws, "ANIO", "ANO", "AÑO", "YEAR")

    filas: list[FilaMonto] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        regimen = str(row[1]).strip() if row[1] is not None else ""
        nombre = str(row[2]).strip() if row[2] is not None else ""

        if not nombre or not regimen:
            continue

        rfc_val = _celda(row, col_rfc)
        rfc = str(rfc_val).strip().upper() if rfc_val else None

        filas.append(FilaMonto(
            fila=i,
            categoria=str(row[0]).strip() if row[0] else None,
            regimen_fiscal=regimen,
            nombre_emisor=normalizar_nombre(nombre),
            rfc_emisor=rfc,
            mes=_parse_mes(_celda(row, col_mes)),
            anio=_parse_anio(_celda(row, col_anio)),
            subtotal=_dec(row[3]),
            iva_trasladado=_dec(row[4]),
            iva_retenido=_dec(row[5]),
            isr_retenido=_dec(row[6]),
            total=_dec(row[7]),
        ))

    return filas
